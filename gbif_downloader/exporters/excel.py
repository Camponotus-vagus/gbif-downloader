"""
Excel exporter with conditional formatting.

Exports occurrence records to Excel format with:
- Yellow highlighting for records with unknown coordinate uncertainty
- Proper column widths
- Frozen header row
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd

from gbif_downloader.api import OccurrenceRecord
from gbif_downloader.utils import get_logger

# Try to import openpyxl for styling
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelExporter:
    """
    Export occurrence records to Excel format.

    Features:
    - Conditional formatting (yellow for unknown uncertainty)
    - Auto-adjusted column widths
    - Frozen header row
    - Hyperlinks for GBIF URLs

    Example:
        exporter = ExcelExporter()
        exporter.export(records, "output.xlsx", highlight_uncertain=True)
    """

    # Yellow fill for uncertain records
    YELLOW_FILL = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    ) if HAS_OPENPYXL else None

    # Header style
    HEADER_FILL = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    ) if HAS_OPENPYXL else None

    HEADER_FONT = Font(
        color="FFFFFF", bold=True
    ) if HAS_OPENPYXL else None

    def __init__(self):
        """Initialize the exporter."""
        self.logger = get_logger()

    def export(
        self,
        records: list[OccurrenceRecord],
        output_path: str | Path,
        highlight_uncertain: bool = True,
    ) -> Path:
        """
        Export records to Excel file.

        Args:
            records: List of OccurrenceRecord objects
            output_path: Output file path
            highlight_uncertain: Highlight rows with unknown uncertainty

        Returns:
            Path to the created file
        """
        output_path = Path(output_path)

        # Ensure .xlsx extension
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")

        self.logger.info(f"Exporting {len(records):,} records to Excel...")

        # Convert records to dictionaries
        data = [record.to_dict() for record in records]
        df = pd.DataFrame(data)

        if not HAS_OPENPYXL or not highlight_uncertain:
            # Simple export without styling
            df.to_excel(output_path, index=False, engine="openpyxl")
            self.logger.info(f"Excel file saved: {output_path}")
            return output_path

        # Export with styling
        self._export_with_styling(df, output_path, highlight_uncertain)

        return output_path

    def _export_with_styling(
        self,
        df: pd.DataFrame,
        output_path: Path,
        highlight_uncertain: bool,
    ) -> None:
        """
        Export with conditional formatting and styling.

        Args:
            df: DataFrame to export
            output_path: Output file path
            highlight_uncertain: Highlight rows with unknown uncertainty
        """
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GBIF Data"

        # Write header
        for col_idx, column in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Find uncertainty column index
        unc_col_idx = None
        for idx, col in enumerate(df.columns):
            if "uncertainty" in col.lower():
                unc_col_idx = idx
                break

        # Write data rows
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            is_uncertain = False

            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Check if this is the uncertainty column
                if col_idx - 1 == unc_col_idx:
                    if pd.isna(value) or value is None or value == "":
                        is_uncertain = True

                # Make links clickable
                if isinstance(value, str) and value.startswith("http"):
                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")

            # Apply yellow highlight to entire row if uncertain
            if highlight_uncertain and is_uncertain:
                for col_idx in range(1, len(row) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = self.YELLOW_FILL

        # Auto-adjust column widths
        for col_idx, column in enumerate(df.columns, start=1):
            max_length = len(str(column))

            # Sample first 100 rows for width calculation
            for row in df.head(100).itertuples(index=False):
                try:
                    cell_value = row[col_idx - 1]
                    if cell_value is not None:
                        max_length = max(max_length, len(str(cell_value)))
                except (IndexError, TypeError):
                    pass

            # Set width with some padding, max 50 characters
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = adjusted_width

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Save
        wb.save(output_path)
        self.logger.info(f"Excel file saved with styling: {output_path}")

    @staticmethod
    def is_available() -> bool:
        """Check if Excel export with styling is available."""
        return HAS_OPENPYXL
